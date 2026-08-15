#!/usr/bin/env python3
"""Lê os 5 arquivos reais da Marquesa e gera seed.sql — os dados iniciais
para o banco D1 na nuvem, já passando pelas mesmas regras que o dashboard
aplicaria numa importação pela tela (consolidação de sufixo, preço
congelado, movimentações), só que sem precisar abrir navegador nenhum.

Cada mudança de estoque é gravada como INSERT + registro em `movimentos`
+ UPDATE do saldo — exatamente a sequência que a API executa em
POST /api/produtos/importar e POST /api/maletas/:id/itens — para o banco
nascer já obedecendo §19 (saldo == soma dos movimentos) sem um caminho
paralelo que possa divergir do código real.
"""
import csv, re, json, collections, sys
import openpyxl

import glob, os

# Caminhos configuráveis por variável de ambiente — nada de nome real de
# pessoa ou caminho específico de máquina fica escrito neste arquivo, já
# que ele é versionado num repositório público. O nome de cada
# revendedora vem de dentro do próprio Anexo (ler_anexo), não daqui.
PASTA = os.environ.get("MARQUESA_DADOS", ".")
F_ESTOQUE = os.environ.get("MARQUESA_ESTOQUE") or next(iter(glob.glob(PASTA + "/*Estoque*.xlsx")), None)
F_LOJA    = os.environ.get("MARQUESA_LOJA")    or next(iter(glob.glob(PASTA + "/*.csv")), None)
ANEXOS    = sorted(glob.glob(PASTA + "/*Anexo*.xlsx"))

if not F_ESTOQUE or not F_LOJA or not ANEXOS:
    raise SystemExit(
        "Não encontrei os arquivos. Rode assim:\n"
        "  MARQUESA_DADOS=/caminho/da/pasta python3 gerar-seed.py\n"
        "A pasta deve conter: a planilha de Estoque (.xlsx), o export da\n"
        "Nuvemshop (.csv) e um ou mais Anexo I (.xlsx)."
    )

CATEGORIAS_VALIDAS = {"Colar","Brinco","Pulseira","Berloque","Anel","Argola","Pingente","Conjunto","Outros"}
CAT_MAP = {
    "colar":"Colar","colares":"Colar","corrente":"Colar","cordao":"Colar","choker":"Colar",
    "gargantilha":"Colar","escapulario":"Colar",
    "brinco":"Brinco","brincos":"Brinco","binco":"Brinco","bincos":"Brinco","trio":"Brinco","duo":"Brinco",
    "dupla":"Brinco","piercing":"Brinco","piecing":"Brinco",
    "argola":"Argola","argolas":"Argola",
    "pulseira":"Pulseira","pulseiras":"Pulseira","bracelete":"Pulseira","tornozeleira":"Pulseira",
    "berloque":"Berloque","separador":"Berloque",
    "anel":"Anel","aneis":"Anel","aparador":"Anel",
    "pingente":"Pingente",
    "conjunto":"Conjunto","conjuntos":"Conjunto",
}

def strip_acentos(s):
    tr = str.maketrans("áàâãéêíóôõúçÁÀÂÃÉÊÍÓÔÕÚÇ", "aaaaeeioooucAAAAEEIOOOUC")
    return s.translate(tr)

def cat_of(desc):
    w = strip_acentos(str(desc or "").strip().lower()).split()
    w = re.sub(r"[^a-z]", "", w[0]) if w else ""
    return CAT_MAP.get(w, "Outros")

def base_sku(s):
    return re.sub(r"-\d+$", "", str(s or "").strip().upper())

def norm_sku(s):
    return str(s or "").strip().upper()

def money(v):
    try: return round(float(str(v).replace(",", ".")), 2)
    except: return 0.0

def sqlstr(v):
    if v is None: return "NULL"
    return "'" + str(v).replace("'", "''") + "'"

def sqlnum(v):
    return "NULL" if v is None else str(v)


# ============================================================ 1. catálogo
def ler_catalogo():
    wb = openpyxl.load_workbook(F_ESTOQUE, data_only=True)["Estoque"]
    linhas = []
    for r in wb.iter_rows(min_row=2, values_only=True):
        if r[0] is None: continue
        bruto = norm_sku(r[0])
        linhas.append({
            "bruto": bruto, "sku": base_sku(bruto),
            "desc": str(r[1] or "").strip(),
            "qtd": int(float(str(r[2]).replace(",", "."))) if r[2] is not None else 0,
            "preco": money(r[3]) if r[3] is not None else 0.0,
        })

    juntos, somados, conflitos = {}, [], []
    for l in linhas:
        ex = juntos.get(l["sku"])
        if ex is None:
            juntos[l["sku"]] = dict(l)
            continue
        a, b = ex["desc"].strip().lower(), l["desc"].strip().lower()
        if a and b and a != b:
            conflitos.append((l["sku"], ex["desc"], l["desc"]))
        ex["qtd"] += l["qtd"]
        if not ex["preco"] and l["preco"]:
            ex["preco"] = l["preco"]
        if l["bruto"] != l["sku"]:
            somados.append(l["bruto"])

    produtos = {}
    for sku, l in juntos.items():
        cat = cat_of(l["desc"])
        produtos[sku] = {
            "sku": sku, "desc": l["desc"] or sku, "cat": cat,
            "preco": l["preco"] if l["preco"] else None,   # §24: 0 vira "sem preço"
            "qtd": l["qtd"],
        }
    return produtos, somados, conflitos


# ================================================================ 2. loja
def ler_loja(produtos):
    raw = open(F_LOJA, encoding="latin-1").read()
    linhas = list(csv.DictReader(raw.splitlines(), delimiter=";"))

    urls_arquivo = set()
    por_sku = {}          # base_sku(CSV) -> {urlLoja, estoqueLoja, visivel, nomeLoja}
    urls_por_sku = collections.defaultdict(set)

    for row in linhas:
        url = (row.get("Identificador URL") or "").strip()
        sku = base_sku(row.get("SKU") or "")
        if url:
            urls_arquivo.add(url)
        if not sku or not url:
            continue
        urls_por_sku[sku].add(url)
        est = 0
        try: est = int(float(str(row.get("Estoque") or "0").replace(",", ".")))
        except: pass
        vis_txt = (row.get("Visibilidade") or "").strip()
        vis = (True if "vis" in vis_txt.lower() else False) if vis_txt else None
        if sku in por_sku:
            por_sku[sku]["estoqueLoja"] += est
            if not por_sku[sku]["urlLoja"]:
                por_sku[sku]["urlLoja"] = url
            if por_sku[sku]["visivel"] is None and vis is not None:
                por_sku[sku]["visivel"] = vis
        else:
            por_sku[sku] = {"urlLoja": url, "estoqueLoja": est, "visivel": vis,
                             "nomeLoja": (row.get("Nome") or "").strip()}

    casados = {s: v for s, v in por_sku.items() if s in produtos}
    urls_casadas = set()
    for s in casados: urls_casadas |= urls_por_sku[s]

    duplicados = sorted(s for s, u in urls_por_sku.items() if len(u) > 1 and s in produtos)

    snapshot = {
        "produtosNaLoja": len(urls_arquivo),
        "produtosCasados": len(urls_casadas),
        "soNaLoja": len(urls_arquivo) - len(urls_casadas),
        "codigosCasados": len(casados),
        "duplicados": duplicados,
    }
    return casados, snapshot


# =============================================================== 3. anexos
def ler_anexo(caminho):
    ws = openpyxl.load_workbook(caminho, data_only=True)["Planilha1"]
    linhas = [list(r) for r in ws.iter_rows(values_only=True)]
    hi = next(i for i, r in enumerate(linhas) if r[1] == "Código")

    # o nome da revendedora é a 2ª linha do documento (logo abaixo do
    # título "Anexo I do contrato") — vem do arquivo, nunca do script
    nome = None
    for r in linhas[:hi]:
        texto = next((str(c).strip() for c in r if c), None)
        if texto and "anexo" not in texto.lower():
            nome = texto; break

    data = None
    for r in linhas[:hi]:
        for c in r:
            m = re.search(r"(\d{2})/(\d{2})/(\d{4})", str(c or ""))
            if m:
                data = f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
                break
        if data: break

    itens = collections.Counter()
    val_doc = {}
    for r in linhas[hi + 1:]:
        if r[1] is None: continue
        sku = base_sku(r[1])
        itens[sku] += 1
        val_doc[sku] = money(r[3])
    return nome, data, itens, val_doc


# ================================================================== main
def main():
    produtos, somados, conflitos = ler_catalogo()
    print(f"catálogo: {len(produtos)} códigos (consolidação somou {len(somados)}: {somados})")
    if conflitos:
        print(f"  ATENÇÃO conflitos de descrição: {conflitos}")
    sem_preco = [s for s, p in produtos.items() if p["preco"] is None]
    print(f"  {len(sem_preco)} sem preço: {sem_preco}")

    casados, snapshot = ler_loja(produtos)
    print(f"loja: {snapshot}")

    maletas = []
    briga_preco = []
    for caminho in ANEXOS:
        nome, acerto_em, itens, val_doc = ler_anexo(caminho)
        if not nome:
            print(f"  ATENÇÃO: não achei o nome da revendedora em {caminho} — pulado")
            continue
        for sku, qtd in itens.items():
            if sku not in produtos:
                print(f"  ATENÇÃO {nome}: código {sku} não está no catálogo — pulado")
                continue
            catp = produtos[sku]["preco"]
            docp = val_doc.get(sku)
            if catp and docp and abs(catp - docp) > 0.005:
                briga_preco.append((nome, sku, docp, catp))
        maletas.append({"revendedora": nome, "acertoEm": acerto_em,
                         "itens": {s: q for s, q in itens.items() if s in produtos}})
        total = sum(q for s, q in itens.items() if s in produtos)
        print(f"maleta {nome}: {total} peças, acerto em {acerto_em}")
    if briga_preco:
        print(f"  preços diferentes entre Anexo e catálogo (uso sempre o catálogo): {briga_preco}")

    # ---------------------------------------------------------------- SQL
    out = []
    out.append("-- Seed gerado a partir dos arquivos reais da Marquesa.")
    out.append("-- Roda em cima de um banco já com schema.sql aplicado (categorias inclusas).")
    out.append("--")
    out.append("-- De propósito, este arquivo não abre nem fecha um bloco de transação: o")
    out.append("-- D1 remoto recusa esse controle vindo de um arquivo executado via --file")
    out.append("-- (ele já trata cada instrução/lote como atômico por conta própria). O")
    out.append("-- wrangler também faz uma busca de texto simples por esse comando — até um")
    out.append("-- COMENTÁRIO citando essas duas palavras juntas dispara o mesmo erro, então")
    out.append("-- evite escrevê-las aqui por extenso, mesmo como explicação.")
    out.append("")

    out.append("-- ============ catálogo ============")
    for sku, p in sorted(produtos.items()):
        cat = p["cat"] if p["cat"] in CATEGORIAS_VALIDAS else "Outros"
        out.append(
            f"INSERT INTO produtos (sku, desc, cat, preco, qtd) VALUES "
            f"({sqlstr(sku)}, {sqlstr(p['desc'])}, {sqlstr(cat)}, {sqlnum(p['preco'])}, 0);"
        )
        out.append(
            f"INSERT INTO movimentos (sku, tipo, qtd, origem, obs) VALUES "
            f"({sqlstr(sku)}, 'entrada', {p['qtd']}, 'importacao', 'Saldo inicial da importação');"
        )
        out.append(f"UPDATE produtos SET qtd = qtd + {p['qtd']} WHERE sku = {sqlstr(sku)};")
    out.append("")

    out.append("-- ============ retrato da Nuvemshop ============")
    out.append(
        "INSERT INTO loja_snapshot (id, lido_em, produtos_na_loja, produtos_casados, so_na_loja, codigos_casados, duplicados_json) "
        f"VALUES (1, datetime('now'), {snapshot['produtosNaLoja']}, {snapshot['produtosCasados']}, "
        f"{snapshot['soNaLoja']}, {snapshot['codigosCasados']}, {sqlstr(json.dumps(snapshot['duplicados']))});"
    )
    for sku, v in sorted(casados.items()):
        vis_sql = "NULL" if v["visivel"] is None else ("1" if v["visivel"] else "0")
        out.append(
            f"UPDATE produtos SET url_loja={sqlstr(v['urlLoja'])}, estoque_loja={v['estoqueLoja']}, "
            f"visivel={vis_sql}, nome_loja={sqlstr(v['nomeLoja'])} WHERE sku={sqlstr(sku)};"
        )
    out.append("")

    out.append("-- ============ revendedoras e maletas ============")
    for m in maletas:
        out.append(f"INSERT INTO revendedoras (nome, status) VALUES ({sqlstr(m['revendedora'])}, 'ativa');")
        out.append(
            f"INSERT INTO maletas (rev_id, status, aberta_em, acerto_em) VALUES "
            f"(last_insert_rowid(), 'aberta', NULL, {sqlstr(m['acertoEm'])});"
        )
        for sku, qtd in sorted(m["itens"].items()):
            preco = produtos[sku]["preco"]
            out.append(
                f"INSERT INTO maleta_itens (maleta_id, sku, qtd, preco_envio) VALUES "
                f"((SELECT id FROM maletas WHERE rev_id = (SELECT id FROM revendedoras WHERE nome = {sqlstr(m['revendedora'])}) "
                f"ORDER BY id DESC LIMIT 1), {sqlstr(sku)}, {qtd}, {sqlnum(preco)});"
            )
            out.append(
                f"INSERT INTO movimentos (sku, tipo, qtd, origem, maleta_id, revendedora_id, obs) VALUES "
                f"({sqlstr(sku)}, 'consignacao', 0, 'maleta', "
                f"(SELECT id FROM maletas WHERE rev_id = (SELECT id FROM revendedoras WHERE nome = {sqlstr(m['revendedora'])}) ORDER BY id DESC LIMIT 1), "
                f"(SELECT id FROM revendedoras WHERE nome = {sqlstr(m['revendedora'])}), "
                f"'Maleta montada na importação inicial');"
            )
    out.append("")

    out.append("-- ============ configuração ============")
    out.append("INSERT OR REPLACE INTO config (chave, valor) VALUES ('prazoDias', '45');")
    out.append("INSERT OR REPLACE INTO config (chave, valor) VALUES ('prataPct', '10');")
    faixas = json.dumps([{"limite":800,"pct":0},{"limite":1000,"pct":15},{"limite":1500,"pct":25},
                          {"limite":6000,"pct":30},{"limite":10000,"pct":35},{"limite":None,"pct":40}])
    out.append(f"INSERT OR REPLACE INTO config (chave, valor) VALUES ('faixas', {sqlstr(faixas)});")

    with open("seed.sql", "w") as f:
        f.write("\n".join(out) + "\n")
    print(f"\nseed.sql gerado: {len(out)} linhas")

if __name__ == "__main__":
    main()
